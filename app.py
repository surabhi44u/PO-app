import io
import re
from typing import Dict

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ----------------------------------
# Streamlit UI
# ----------------------------------
st.set_page_config(page_title="PO Generator (Fixed Mapping)", page_icon="📄", layout="wide")
st.title("📄 Purchase Order Generator — Fixed Mapping")
st.caption("Green areas stay static. Blue fields are filled from your input. One sheet per (Control NO, Item NO). If auto-detect fails, use the mapping dropdowns.")

with st.sidebar:
    st.header("Inputs")
    input_file = st.file_uploader("Upload INPUT Excel (.xlsm/.xlsx)", type=["xlsm", "xlsx"])
    sheet_name = st.text_input("Input sheet name", value="250826")
    header_row_1based = st.number_input(
        "Header row (1 = first row)", min_value=1, value=1, step=1,
        help="If your headers aren't on the first row, set the correct row here."
    )
    template_file = st.file_uploader("Upload TEMPLATE .xlsx", type=["xlsx"], help="Use the colored template you provided")
    remove_template_sheet = st.checkbox("Remove the original template sheet from output", value=True)
    show_preview = st.checkbox("Show input preview & detected columns", value=True)
    btn = st.button("Generate Purchase Orders")

# ----------------------------------
# Helpers
# ----------------------------------
INVALID_SHEET_CHARS = r"[:\\\\/?*\[\]]"

FIXED_CELL_MAP = {
    "control_no": "AD9",
    "item_no": "E16",
    "barcode": "S16",
    "delivery": "B28",
    "qty": "AA24",
    "price1": "N30",
    "price2": "N32",
    "amount": "F37",
}

CELLS_TO_CLEAR = ["E18", "E20", "E24", "N24", "B26", "A35", "R37", "F39"]


def sanitize_sheet_title(title: str) -> str:
    title = re.sub(INVALID_SHEET_CHARS, "-", str(title)).strip() or "Sheet"
    return title[:31]

# Likely header names in your files (we'll try these first)
PREFERRED = {
    "control_no": ["Control NO", "CONTROL NO", "Control No", "control no", "ControlNO", "Ctrl No", "Control"],
    "item_no": ["Item NO", "ITEM NO", "Item No", "item no", "Item code", "品番", "品番 / Item no"],
    "barcode": ["Barcode", "JAN", "JAN code", "JAN Code", "JANコード"],
    "qty": ["Qty", "QTY", "Quantity", "数量"],
    "price": ["Price", "単価", "Unit Price", "Unit price"],
    "delivery": ["Delivery time", "Delivery", "Delivery date", "納期"],
}


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = (
        pd.Series(df.columns, dtype="string")
        .fillna("")
        .str.replace("\u00A0", " ", regex=False)
        .str.replace("\n", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    df.columns = cols
    return df


@st.cache_data(show_spinner=False)
def load_input(_file, _sheet, header_idx: int) -> pd.DataFrame:
    df = pd.read_excel(_file, sheet_name=_sheet, engine="openpyxl", header=header_idx)
    return normalize_columns(df)


def find_col(df: pd.DataFrame, candidates) -> str:
    # exact
    for cand in candidates:
        for c in df.columns:
            if c == cand:
                return c
    # case-insensitive exact
    for cand in candidates:
        cl = cand.lower()
        for c in df.columns:
            if c.lower() == cl:
                return c
    # contains
    for cand in candidates:
        cl = cand.lower()
        for c in df.columns:
            if cl in c.lower():
                return c
    return None


# ----------------------------------
# Main action
# ----------------------------------
if btn:
    if not input_file or not template_file:
        st.error("Please upload both the INPUT workbook and the TEMPLATE workbook.")
        st.stop()

    try:
        df = load_input(input_file, sheet_name, header_row_1based - 1)
    except Exception as e:
        st.error(f"Could not read sheet '{sheet_name}': {e}")
        st.stop()

    if df.empty:
        st.error("The input sheet appears to be empty.")
        st.stop()

    if show_preview:
        st.subheader("Input preview")
        st.dataframe(df.head(20))
        st.write("Detected headers:", list(df.columns))

    # Auto-detect columns
    cols: Dict[str, str] = {}
    for key, cands in PREFERRED.items():
        cols[key] = find_col(df, cands)

    required = ["control_no", "item_no", "barcode", "qty", "price", "delivery"]
    missing = [k for k in required if not cols.get(k)]

    # Fallback mapping UI if any missing
    if missing:
        st.warning("Auto-detection failed for some fields. Please map them manually.")
        for key in required:
            guess = cols.get(key)
            cols[key] = st.selectbox(
                f"Select column for {key.replace('_',' ').title()}",
                options=[guess] + [c for c in df.columns if c != guess] if guess else [None] + list(df.columns),
                index=0,
            )
        if any(v is None for v in cols.values()):
            st.stop()

    # Deduplicate: first row per (Control NO, Item NO)
    df_sorted = df.copy()
    df_sorted["__group_key__"] = (
        df_sorted[cols["control_no"]].astype(str).str.strip() + "\u0001" +
        df_sorted[cols["item_no"]].astype(str).str.strip()
    )
    first_rows = df_sorted.drop_duplicates("__group_key__", keep="first").reset_index(drop=True)

    def to_float(x):
        if pd.isna(x):
            return None
        s = str(x)
        s = s.replace(",", "").replace("￥", "").replace("¥", "").strip()
        try:
            return float(s)
        except Exception:
            try:
                return float(s.replace(" ", ""))
            except Exception:
                return None

    def to_int(x):
        f = to_float(x)
        return int(round(f)) if f is not None else None

    # Load template workbook
    try:
        wb = load_workbook(template_file, data_only=False, keep_vba=False)
    except Exception as e:
        st.error(f"Failed to open template: {e}")
        st.stop()

    tpl_ws = wb.active
    created = []

    for _, r in first_rows.iterrows():
        control_no = str(r[cols["control_no"]]).strip()
        item_no    = str(r[cols["item_no"]]).strip()
        barcode    = str(r[cols["barcode"]]).strip()
        qty        = to_int(r[cols["qty"]])
        price      = to_float(r[cols["price"]])
        delivery   = str(r[cols["delivery"]]).strip()

        ws = wb.copy_worksheet(tpl_ws)
        ws.title = sanitize_sheet_title(f"{control_no}_{item_no}")

        # Fill dynamic cells (BLUE)
        try: ws[FIXED_CELL_MAP["control_no"]] = control_no
        except: pass
        try: ws[FIXED_CELL_MAP["item_no"]] = item_no
        except: pass
        try: ws[FIXED_CELL_MAP["barcode"]] = barcode
        except: pass
        try: ws[FIXED_CELL_MAP["delivery"]] = delivery
        except: pass
        try: ws[FIXED_CELL_MAP["qty"]] = qty
        except: pass
        try: ws[FIXED_CELL_MAP["price1"]] = price
        except: pass
        try: ws[FIXED_CELL_MAP["price2"]] = price
        except: pass
        try:
            amount = (qty or 0) * (price or 0)
            ws[FIXED_CELL_MAP["amount"]] = amount
        except: pass

        # Clear cells
        for cell in CELLS_TO_CLEAR:
            try: ws[cell] = None
            except: pass

        # Delete rows 60–64
        try: ws.delete_rows(60, 5)
        except: pass

        created.append(ws.title)

    if remove_template_sheet:
        try: wb.remove(tpl_ws)
        except: pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    st.success(f"Created {len(created)} sheet(s)")
    st.download_button(
        "⬇️ Download PurchaseOrders.xlsx",
        buf,
        file_name="PurchaseOrders.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    with st.expander("Detected column mapping"):
        st.json(cols)
    with st.expander("Sheets created"):
        st.write(created)

st.markdown("---")
st.markdown(
    """
**Fixed cell mapping (blue fields):**  AD9 (Control NO), E16 (Item No), S16 (JAN/Barcode), B28 (Delivery), AA24 (Qty), N30 & N32 (Price), F37 (Amount = Qty×Price)

**Cleanup applied to each sheet:**  Clears E18, E20, E24, N24, B26, A35, R37, F39; deletes rows 60–64.

If you still see NULLs, set the **Header row** correctly and, if needed, use the **manual mapping** dropdowns. The preview shows the first 20 rows and the detected headers for troubleshooting.
"""
)
